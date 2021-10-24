import openpyxl

wb=openpyxl.load_workbook("..//excel//testexceldata.xlsx")
sheet=wb["testdata"]
total_rows=sheet.max_row
total_cols=sheet.max_column

print("The maximum rows are ",str(total_rows))
print("The maximum cols are ",str(total_cols))

for rows in range(1,total_rows+1):
    for cols in range(1,total_cols+1):
        print(sheet.cell(row=rows,column=cols).value,end="   ")
    print()
